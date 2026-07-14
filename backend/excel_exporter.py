"""
excel_exporter.py
-----------------
Generates a multi-sheet Excel workbook from a completed analysis result.

Sheets
------
  1. Summary                  — executive overview, key metrics
  2. Recommendations          — every detected recommendation with costs and fix commands
  3. Cost Breakdown           — month-to-date actual spend per resource
  4. General Recommendations  — general improvement recommendations

Usage (standalone / pipeline)
------------------------------
  from excel_exporter import build_excel_bytes
  xlsx_bytes = build_excel_bytes(analysis_result, resource_group, resource_count)
  Path("report.xlsx").write_bytes(xlsx_bytes)
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette (Azure-ish dark-accent)
# ---------------------------------------------------------------------------

_AZURE_BLUE   = "0078D4"
_DARK_HEADER  = "1E2A38"
_WHITE        = "FFFFFF"
_LIGHT_ROW    = "F5F8FC"
_ALT_ROW      = "EAF1FB"

_RED_BG       = "FDE8E8"
_YELLOW_BG    = "FEF9E7"
_GREEN_BG     = "E8F8F0"

_RED_TEXT     = "C0392B"
_YELLOW_TEXT  = "9A6700"
_GREEN_TEXT   = "1E7A45"
_SAVINGS_TEXT = "0D5C30"

_THIN_SIDE = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)


# ---------------------------------------------------------------------------
# Helper: styled header cell
# ---------------------------------------------------------------------------

def _hdr(ws, row: int, col: int, value: str, *, wide: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=_DARK_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _val(ws, row: int, col: int, value: Any, *, bold: bool = False,
         color: str | None = None, bg: str | None = None,
         align: str = "left", wrap: bool = False, fmt: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color or "000000", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _THIN_BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt


def _auto_width(ws, min_w: int = 12, max_w: int = 55) -> None:
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(
            min_w, min(length + 4, max_w)
        )


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_summary(wb: openpyxl.Workbook, analysis: dict, resource_group: str,
                   subscription_id: str | None, resource_count: int | None,
                   generated_at: str) -> None:
    ws = wb.create_sheet("Summary")

    # Banner
    ws.merge_cells("A1:F1")
    banner = ws["A1"]
    banner.value = "AI Cloud Cost Detective — Analysis Report"
    banner.font = Font(bold=True, color=_WHITE, size=14)
    banner.fill = PatternFill("solid", fgColor=_AZURE_BLUE)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Meta info
    meta_rows = [
        ("Resource Group", resource_group),
        ("Subscription", subscription_id or "—"),
        ("Generated At", generated_at),
        ("Resources Scanned", resource_count or "—"),
    ]
    for r, (label, val) in enumerate(meta_rows, start=2):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val)
    ws.row_dimensions[2].height = 18

    # Metrics row
    issues = analysis.get("issues", [])
    high_c   = sum(1 for i in issues if i.get("severity") == "high")
    med_c    = sum(1 for i in issues if i.get("severity") == "medium")
    low_c    = sum(1 for i in issues if i.get("severity") == "low")
    savings  = analysis.get("total_estimated_monthly_savings_usd") or 0.0
    total_cb = sum(e.get("cost_usd", 0) for e in analysis.get("actual_cost_breakdown", []))

    metric_row = 7
    ws.merge_cells(f"A{metric_row}:F{metric_row}")
    ws[f"A{metric_row}"] = "Key Metrics"
    ws[f"A{metric_row}"].font = Font(bold=True, size=11)
    ws.row_dimensions[metric_row].height = 20

    headers  = ["Total Recommendations", "High Severity", "Medium Severity", "Low Severity",
                "Est. Monthly Savings (USD)", "MTD Actual Spend (USD)"]
    values   = [len(issues), high_c, med_c, low_c,
                f"${savings:,.2f}", f"${total_cb:,.2f}"]
    bgs      = [_LIGHT_ROW, _RED_BG, _YELLOW_BG, _GREEN_BG, _GREEN_BG, _LIGHT_ROW]
    texts    = ["000000", _RED_TEXT, _YELLOW_TEXT, _GREEN_TEXT, _SAVINGS_TEXT, "000000"]

    for c, (h, v, bg, txt) in enumerate(zip(headers, values, bgs, texts), start=1):
        _hdr(ws, metric_row + 1, c, h)
        _val(ws, metric_row + 2, c, v, bold=True, bg=bg, color=txt, align="center")

    ws.row_dimensions[metric_row + 2].height = 22

    # Executive summary text
    summary_row = metric_row + 4
    ws.merge_cells(f"A{summary_row}:F{summary_row}")
    ws[f"A{summary_row}"] = "Executive Summary"
    ws[f"A{summary_row}"].font = Font(bold=True, size=11)

    text_row = summary_row + 1
    ws.merge_cells(f"A{text_row}:F{text_row + 5}")
    cell = ws[f"A{text_row}"]
    cell.value = analysis.get("summary", "")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[text_row].height = 100

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.column_dimensions["A"].width = 26


def _sheet_issues(wb: openpyxl.Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("Recommendations")
    ws.freeze_panes = "A2"

    headers = [
        "Resource Name", "Resource Type", "Severity", "Category",
        "Recommendation", "Current Cost/mo (USD)", "Optimized Cost/mo (USD)",
        "Est. Savings/mo (USD)", "Savings Reasoning", "Fix Commands",
    ]
    for c, h in enumerate(headers, start=1):
        _hdr(ws, 1, c, h)
    ws.row_dimensions[1].height = 30

    sev_bg = {"high": _RED_BG, "medium": _YELLOW_BG, "low": _GREEN_BG}
    sev_txt = {"high": _RED_TEXT, "medium": _YELLOW_TEXT, "low": _GREEN_TEXT}

    for r, issue in enumerate(analysis.get("issues", []), start=2):
        sev = (issue.get("severity") or "low").lower()
        bg  = sev_bg.get(sev, _LIGHT_ROW)
        txt = sev_txt.get(sev, "000000")
        alt = _ALT_ROW if r % 2 == 0 else _LIGHT_ROW

        _val(ws, r, 1, issue.get("resource_name", ""), bg=alt)
        _val(ws, r, 2, issue.get("resource_type", ""), bg=alt)
        _val(ws, r, 3, sev.capitalize(), bold=True, bg=bg, color=txt, align="center")
        _val(ws, r, 4, issue.get("category", ""), bg=alt)
        _val(ws, r, 5, issue.get("issue", ""), bg=alt, wrap=True)

        cur  = issue.get("current_monthly_cost_usd")
        opt  = issue.get("optimized_monthly_cost_usd")
        sav  = issue.get("estimated_monthly_savings_usd")

        _val(ws, r, 6, cur if cur is not None else "", bg=alt, align="right",
             fmt='"$"#,##0.00' if cur is not None else None)
        _val(ws, r, 7, opt if opt is not None else "", bg=alt, align="right",
             fmt='"$"#,##0.00' if opt is not None else None)
        _val(ws, r, 8, sav if sav is not None else "", bold=True, bg=alt,
             color=_SAVINGS_TEXT if sav and sav > 0 else "000000",
             align="right", fmt='"$"#,##0.00' if sav is not None else None)
        _val(ws, r, 9, issue.get("savings_reasoning", ""), bg=alt, wrap=True)

        cmds = issue.get("fix_commands", [])
        _val(ws, r, 10, "\n".join(cmds), bg=alt, wrap=True)

        ws.row_dimensions[r].height = max(30, 15 * len(cmds) if cmds else 18)

    _auto_width(ws)
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 55


def _sheet_cost_breakdown(wb: openpyxl.Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("Cost Breakdown")
    ws.freeze_panes = "A2"

    headers = ["Resource Name", "Resource Type", "MTD Cost (USD)", "Original Amount", "Currency"]
    for c, h in enumerate(headers, start=1):
        _hdr(ws, 1, c, h)
    ws.row_dimensions[1].height = 28

    entries = sorted(
        analysis.get("actual_cost_breakdown", []),
        key=lambda x: x.get("cost_usd", 0),
        reverse=True,
    )

    for r, entry in enumerate(entries, start=2):
        bg = _ALT_ROW if r % 2 == 0 else _LIGHT_ROW
        rid   = entry.get("resource_id", "")
        short = rid.split("/")[-1] if rid else ""
        rtype = entry.get("resource_type", "").split("/")[-1]
        cost  = entry.get("cost_usd", 0)

        cost_txt = _RED_TEXT if cost > 10 else (_YELLOW_TEXT if cost > 1 else "000000")

        _val(ws, r, 1, short or rid, bg=bg)
        _val(ws, r, 2, rtype, bg=bg)
        _val(ws, r, 3, cost, bold=True, bg=bg, color=cost_txt,
             align="right", fmt='"$"#,##0.00')
        _val(ws, r, 4, entry.get("cost_original", ""), bg=bg, align="right",
             fmt='#,##0.00')
        _val(ws, r, 5, entry.get("currency_original", "USD"), bg=bg, align="center")

    # Total row
    total_row = len(entries) + 2
    ws.merge_cells(f"A{total_row}:B{total_row}")
    ws.cell(row=total_row, column=1, value="Total this month").font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=_DARK_HEADER)
    ws.cell(row=total_row, column=1).font = Font(bold=True, color=_WHITE, size=10)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    total_val = sum(e.get("cost_usd", 0) for e in entries)
    _val(ws, total_row, 3, total_val, bold=True, bg=_DARK_HEADER,
         color=_WHITE, align="right", fmt='"$"#,##0.00')

    _auto_width(ws)
    ws.column_dimensions["A"].width = 40


def _sheet_recommendations(wb: openpyxl.Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("General Recommendations")

    ws.merge_cells("A1:C1")
    ws["A1"] = "General Recommendations"
    ws["A1"].font = Font(bold=True, color=_WHITE, size=12)
    ws["A1"].fill = PatternFill("solid", fgColor=_AZURE_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for r, rec in enumerate(analysis.get("general_recommendations", []), start=2):
        bg = _ALT_ROW if r % 2 == 0 else _LIGHT_ROW
        ws.cell(row=r, column=1, value=r - 1).font = Font(bold=True, size=10)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

        ws.merge_cells(f"B{r}:C{r}")
        cell = ws.cell(row=r, column=2, value=rec)
        cell.font = Font(size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[r].height = 50

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 10


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_excel_bytes(
    analysis_result: dict[str, Any],
    resource_group: str,
    resource_count: int | None = None,
    subscription_id: str | None = None,
) -> bytes:
    """
    Build an .xlsx report in memory and return the raw bytes.

    Parameters
    ----------
    analysis_result : dict
        The full analysis JSON as stored in `analyses.analysis_result`.
    resource_group : str
        Display name of the scanned resource group.
    resource_count : int | None
        Number of resources scanned (shown in Summary sheet).
    subscription_id : str | None
        Azure subscription ID (shown in Summary sheet).

    Returns
    -------
    bytes
        Raw .xlsx file content ready to stream or write to disk.
    """
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _sheet_summary(wb, analysis_result, resource_group, subscription_id,
                   resource_count, generated_at)
    _sheet_issues(wb, analysis_result)
    _sheet_cost_breakdown(wb, analysis_result)
    _sheet_recommendations(wb, analysis_result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
